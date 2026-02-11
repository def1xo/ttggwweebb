from __future__ import annotations

from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Tuple, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, text
from sqlalchemy.exc import DataError
from sqlalchemy.orm import Session, selectinload

from app.api.dependencies import get_db, get_current_admin_user
from app.db import models

router = APIRouter(tags=["admin_dashboard"])


# Order statuses that count as "sales".
# IMPORTANT: the Postgres enum `order_status` does NOT contain "confirmed".
# Passing an invalid value to .in_(...) causes a DB error.
CONFIRMED_STATUSES: Tuple[str, ...] = (
    models.OrderStatus.paid.value,
    models.OrderStatus.processing.value,
    models.OrderStatus.sent.value,
    models.OrderStatus.received.value,
)


def _db_enum_values(db: Session, enum_name: str) -> Optional[set]:
    """Return PostgreSQL enum labels for a given enum type name.

    If we can't introspect (non-Postgres or insufficient privileges), returns None.
    """
    try:
        rows = db.execute(
            text(
                """
                SELECT e.enumlabel
                FROM pg_type t
                JOIN pg_enum e ON t.oid = e.enumtypid
                WHERE t.typname = :name
                ORDER BY e.enumsortorder
                """
            ),
            {"name": enum_name},
        ).fetchall()
        return {r[0] for r in rows}
    except Exception:
        return None


def _safe_confirmed_statuses(db: Session) -> Tuple[str, ...]:
    """DB may have an older enum set; keep only statuses that exist there."""
    allowed = _db_enum_values(db, "order_status")
    if not allowed:
        return CONFIRMED_STATUSES
    filtered = tuple(s for s in CONFIRMED_STATUSES if s in allowed)
    if filtered:
        return filtered
    # fallback: prefer paid if present, otherwise any available
    if "paid" in allowed:
        return ("paid",)
    return (next(iter(allowed)),)


def _utcnow() -> datetime:
    return datetime.utcnow()


def _range_start(range_name: str) -> datetime | None:
    now = _utcnow()
    rn = (range_name or "").lower().strip()
    if rn in ("week", "7d", "7"):
        return now - timedelta(days=6)
    if rn in ("month", "30d", "30"):
        return now - timedelta(days=29)
    if rn in ("all", "alltime", "all_time"):
        return None
    raise HTTPException(status_code=400, detail="invalid range; use week|month|all")


def _month_start() -> datetime:
    now = _utcnow()
    return now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)


def _latest_cost_map(db: Session, variant_ids: List[int]) -> Dict[int, Decimal]:
    """Return latest known закуп (cost price) per variant.

    Best-effort: if the ProductCost table isn't available yet or any DB error occurs,
    return an empty map so dashboards don't crash.
    """
    if not variant_ids:
        return {}
    try:
        sub = (
            db.query(
                models.ProductCost.variant_id.label("variant_id"),
                func.max(models.ProductCost.created_at).label("max_created"),
            )
            .filter(models.ProductCost.variant_id.in_(variant_ids))
            .group_by(models.ProductCost.variant_id)
            .subquery()
        )
        rows = (
            db.query(models.ProductCost)
            .join(
                sub,
                (models.ProductCost.variant_id == sub.c.variant_id)
                & (models.ProductCost.created_at == sub.c.max_created),
            )
            .all()
        )
    except Exception:
        return {}

    out: Dict[int, Decimal] = {}
    for r in rows:
        try:
            out[int(r.variant_id)] = Decimal(str(getattr(r, "cost_price", 0) or 0))
        except Exception:
            out[int(r.variant_id)] = Decimal("0")
    return out


@router.get("/stats")
def admin_stats(
    range: str = Query("month", description="week|month|all"),
    db: Session = Depends(get_db),
    admin: models.User = Depends(get_current_admin_user),
):
    """Returns sales time-series for chart + current month summary.

    * Chart totals are gross (no commissions/cost deducted).
    """
    statuses = _safe_confirmed_statuses(db)
    start = _range_start(range)

    try:
        q = (
            db.query(
                func.date(models.Order.created_at).label("d"),
                func.coalesce(func.sum(models.Order.total_amount), 0).label("total"),
            )
            .filter(models.Order.status.in_(statuses))
        )
        if start is not None:
            q = q.filter(models.Order.created_at >= start)
        q = q.group_by("d").order_by("d")
        rows = q.all()
    except DataError:
        # Most common cause: database enum is missing one of the statuses.
        statuses = (models.OrderStatus.paid.value,)
        q = (
            db.query(
                func.date(models.Order.created_at).label("d"),
                func.coalesce(func.sum(models.Order.total_amount), 0).label("total"),
            )
            .filter(models.Order.status.in_(statuses))
        )
        if start is not None:
            q = q.filter(models.Order.created_at >= start)
        q = q.group_by("d").order_by("d")
        rows = q.all()
    except Exception:
        logger.exception("Admin stats query failed")
        rows = []

    series: List[Dict[str, object]] = []
    for d, total in rows:
        try:
            ds = d.isoformat()  # type: ignore[attr-defined]
        except Exception:
            ds = str(d)
        series.append({"date": ds, "amount": float(total or 0), "total": float(total or 0)})

    # month summary (calendar month)
    m_start = _month_start()
    try:
        orders = (
            db.query(models.Order)
            .options(selectinload(models.Order.items))
            .filter(models.Order.created_at >= m_start)
            .filter(models.Order.status.in_(statuses))
            .order_by(models.Order.created_at.desc())
            .all()
        )
    except DataError:
        statuses = (models.OrderStatus.paid.value,)
        orders = (
            db.query(models.Order)
            .options(selectinload(models.Order.items))
            .filter(models.Order.created_at >= m_start)
            .filter(models.Order.status.in_(statuses))
            .order_by(models.Order.created_at.desc())
            .all()
        )
    except Exception:
        logger.exception("Admin stats orders fetch failed")
        orders = []

    revenue = Decimal("0")
    for o in orders:
        try:
            revenue += Decimal(str(o.total_amount or 0))
        except Exception:
            revenue += Decimal("0")

    # compute закуп (COGS) from latest ProductCost per variant
    variant_ids_set = set()
    for o in orders:
        for it in (o.items or []):
            if getattr(it, "variant_id", None) is not None:
                variant_ids_set.add(int(it.variant_id))
    cost_map = _latest_cost_map(db, sorted(list(variant_ids_set)))

    cost_total = Decimal("0")
    for o in orders:
        for it in (o.items or []):
            try:
                qty = int(getattr(it, "quantity", 0) or 0)
            except Exception:
                qty = 0
            if qty <= 0:
                continue
            vid = getattr(it, "variant_id", None)
            if vid is None:
                continue
            c = cost_map.get(int(vid), Decimal("0"))
            cost_total += (c * Decimal(qty))

    profit = revenue - cost_total
    margin_percent = float((profit / revenue * Decimal("100")) if revenue > 0 else 0)

    revenue_f = float(revenue)
    cost_f = float(cost_total)
    profit_f = float(profit)


    return {
        "range": (range or "month"),
        "series": series,
        "month": {
            "month_start": m_start.isoformat() + "Z",
            "orders_count": len(orders),
            # keep both legacy and current keys for frontend compatibility
            "revenue": revenue_f,
            "cost": cost_f,
            "profit": profit_f,
            "revenue_gross": revenue_f,
            "cogs_estimated": cost_f,
            "profit_estimated": profit_f,
            "margin_percent": margin_percent,
        },
    }


@router.get("/export/sales.xlsx")
def export_sales_xlsx(
    scope: str = Query("month", description="month|week|all"),
    db: Session = Depends(get_db),
    admin: models.User = Depends(get_current_admin_user),
):
    """Export orders and simple gross/cost metrics to XLSX."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    now = _utcnow()
    sc = (scope or "month").lower().strip()
    if sc == "month":
        start = _month_start()
    elif sc == "week":
        start = now - timedelta(days=6)
    elif sc in ("all", "alltime", "all_time"):
        start = None
    else:
        raise HTTPException(status_code=400, detail="invalid scope; use month|week|all")

    statuses = _safe_confirmed_statuses(db)

    def _fetch(sts: Tuple[str, ...]):
        q = (
            db.query(models.Order)
            .options(selectinload(models.Order.items))
            .filter(models.Order.status.in_(sts))
            .order_by(models.Order.created_at.desc())
        )
        if start is not None:
            q = q.filter(models.Order.created_at >= start)
        return q.all()

    try:
        orders = _fetch(statuses)
    except DataError:
        orders = _fetch((models.OrderStatus.paid.value,))
        statuses = (models.OrderStatus.paid.value,)
    except Exception:
        logger.exception("Admin export query failed")
        orders = []

    # cost map
    variant_ids_set = set()
    for o in orders:
        for it in (o.items or []):
            if getattr(it, "variant_id", None) is not None:
                variant_ids_set.add(int(it.variant_id))
    cost_map = _latest_cost_map(db, sorted(list(variant_ids_set)))

    def order_cost(o: models.Order) -> Decimal:
        total = Decimal("0")
        for it in (o.items or []):
            try:
                qty = int(getattr(it, "quantity", 0) or 0)
            except Exception:
                qty = 0
            if qty <= 0:
                continue
            vid = getattr(it, "variant_id", None)
            if vid is None:
                continue
            total += cost_map.get(int(vid), Decimal("0")) * Decimal(qty)
        return total

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    bold = Font(bold=True)

    # compute summary
    revenue = Decimal("0")
    cost_total = Decimal("0")
    for o in orders:
        revenue += Decimal(str(o.total_amount or 0))
        cost_total += order_cost(o)
    profit = revenue - cost_total
    margin_percent = float((profit / revenue * Decimal("100")) if revenue > 0 else 0)

    rows = [
        ("Scope", sc),
        ("Generated", now.isoformat() + "Z"),
        ("Orders count", len(orders)),
        ("Revenue (gross)", float(revenue)),
        ("Cost (zakup)", float(cost_total)),
        ("Profit", float(profit)),
        ("Margin %", margin_percent),
    ]
    for r, (k, v) in enumerate(rows, start=1):
        ws.cell(r, 1, k).font = bold
        ws.cell(r, 2, v)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28

    ws2 = wb.create_sheet("Orders")
    header = ["Order ID", "Created at", "Status", "Buyer", "Revenue", "Cost", "Profit"]
    ws2.append(header)
    for c in range(1, len(header) + 1):
        ws2.cell(1, c).font = bold
        ws2.cell(1, c).alignment = Alignment(horizontal="center")

    for o in orders:
        oc = order_cost(o)
        rev = Decimal(str(o.total_amount or 0))
        ws2.append(
            [
                o.id,
                (o.created_at.isoformat() + "Z") if getattr(o, "created_at", None) else "",
                o.status,
                getattr(o, "fio", None) or "",
                float(rev),
                float(oc),
                float(rev - oc),
            ]
        )

    # basic formatting
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 26
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 22
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 14

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"sales_{sc}_{now.strftime('%Y-%m-%d')}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/ops/needs-attention")
def admin_ops_needs_attention(
    low_stock_threshold: int = Query(2, ge=0, le=50),
    stale_order_hours: int = Query(24, ge=1, le=168),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    admin: models.User = Depends(get_current_admin_user),
):
    """Operational queue for admins: stale orders + problematic products/stock."""
    now = _utcnow()
    stale_cutoff = now - timedelta(hours=stale_order_hours)

    stale_orders = (
        db.query(models.Order)
        .filter(models.Order.status == models.OrderStatus.awaiting_payment.value)
        .filter(models.Order.created_at <= stale_cutoff)
        .order_by(models.Order.created_at.asc())
        .limit(limit)
        .all()
    )

    products = (
        db.query(models.Product)
        .options(selectinload(models.Product.variants))
        .order_by(models.Product.updated_at.desc().nullslast(), models.Product.created_at.desc())
        .limit(800)
        .all()
    )

    missing_cards = []
    low_stock = []

    for p in products:
        variants = list(p.variants or [])
        has_variants = len(variants) > 0
        base_price = float(p.base_price or 0)
        has_description = bool((p.description or "").strip())
        has_image = bool((p.default_image or "").strip())

        reasons = []
        if not has_description:
            reasons.append("no_description")
        if not has_image:
            reasons.append("no_image")
        if base_price <= 0:
            reasons.append("price_zero")
        if not has_variants:
            reasons.append("no_variants")

        if reasons:
            missing_cards.append(
                {
                    "product_id": p.id,
                    "title": p.title,
                    "visible": bool(p.visible),
                    "reasons": reasons,
                    "updated_at": p.updated_at.isoformat() + "Z" if getattr(p, "updated_at", None) else None,
                }
            )

        for v in variants:
            stock = int(v.stock_quantity or 0)
            if stock <= low_stock_threshold:
                low_stock.append(
                    {
                        "variant_id": v.id,
                        "product_id": p.id,
                        "title": p.title,
                        "stock_quantity": stock,
                        "is_out": stock <= 0,
                    }
                )

    missing_cards = missing_cards[:limit]
    low_stock.sort(key=lambda x: (x["stock_quantity"], x["product_id"]))
    low_stock = low_stock[:limit]

    stale_without_proof = 0

    stale_serialized = [
        {
            "order_id": o.id,
            "created_at": o.created_at.isoformat() + "Z" if getattr(o, "created_at", None) else None,
            "hours_waiting": round(max(0.0, (now - o.created_at).total_seconds()) / 3600, 1) if getattr(o, "created_at", None) else 0,
            "total_amount": float(o.total_amount or 0),
            "fio": o.fio,
            "phone": o.phone,
            "has_payment_proof": bool(o.payment_screenshot),
        }
        for o in stale_orders
    ]

    stale_without_proof = sum(1 for o in stale_serialized if not bool(o.get("has_payment_proof")))
    severity_score = (
        len(stale_serialized) * 3
        + len(missing_cards) * 2
        + len(low_stock) * 1
    )

    return {
        "generated_at": now.isoformat() + "Z",
        "thresholds": {
            "low_stock_threshold": low_stock_threshold,
            "stale_order_hours": stale_order_hours,
        },
        "counts": {
            "stale_orders": len(stale_serialized),
            "stale_orders_without_proof": stale_without_proof,
            "products_missing_data": len(missing_cards),
            "low_stock_variants": len(low_stock),
        },
        "severity_score": severity_score,
        "items": {
            "stale_orders": stale_serialized,
            "products_missing_data": missing_cards,
            "low_stock_variants": low_stock,
        },
    }
